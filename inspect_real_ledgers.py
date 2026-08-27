from openpyxl import load_workbook
path='/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx'
wb=load_workbook(path,data_only=True,read_only=True)
for name in ['CONTROLE- Fluxo de Entrada','CONTROLE - fluxo de saida']:
    print('---',name)
    ws=wb[name]
    shown=0
    for i,row in enumerate(ws.iter_rows(values_only=True),1):
        vals=[v for v in row if v not in (None,'')]
        if vals:
            print(i, vals[:12])
            shown+=1
        if shown>=18: break
