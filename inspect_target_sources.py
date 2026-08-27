from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(path, data_only=False)
for sheet, rows, cols in [('CONTROLE- Fluxo de Entrada', range(1, 22), range(1, 9)), ('Boletos a Pagar- AGOSTO', range(1, 22), range(1, 11))]:
    ws = wb[sheet]
    print('\nSHEET', sheet)
    for r in rows:
        vals = [ws.cell(r, c).value for c in cols]
        if any(v is not None for v in vals): print(r, vals)
