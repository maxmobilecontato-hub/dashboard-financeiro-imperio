from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl import load_workbook
import re
from html import escape

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_boleto_grafico_conectado.xlsx')

# Read cached real data only to populate the initial view; formulas remain dynamic.
data_wb = load_workbook(source, data_only=True, read_only=True)
bills = data_wb['Boletos a Pagar- AGOSTO']
paid = []
for row in range(12, 1001):
    status = bills.cell(row, 1).value
    fallback_status = bills.cell(row, 8).value
    value = bills.cell(row, 7).value
    name = bills.cell(row, 3).value
    if (status == '✓' or str(fallback_status).strip().upper() == 'PAGO') and isinstance(value, (int, float)):
        paid.append((float(value), str(name or '').strip()))
data_wb.close()
paid.sort(key=lambda item: item[0], reverse=True)
top = paid[:5] + [(0.0, '')] * max(0, 5 - len(paid))

with ZipFile(source, 'r') as zin:
    files = {info.filename: zin.read(info.filename) for info in zin.infolist()}

sheet = files['xl/worksheets/sheet1.xml'].decode('utf-8')

def replace_formula_and_cache(xml, cell, formula, cache):
    pattern = rf'<c r="{cell}"([^>]*?)(?:>(.*?)</c>|/>)'
    safe_cache = escape(str(cache), quote=False)
    body = f'<f>{escape(formula, quote=False)}</f><v>{safe_cache}</v>'
    def repl(match):
        attrs = match.group(1)
        return f'<c r="{cell}"{attrs}>{body}</c>'
    new_xml, count = re.subn(pattern, repl, xml, count=1, flags=re.S)
    assert count == 1, cell
    return new_xml

# Repair only the hidden helper rows consumed by the existing Boletos Pagos chart.
for idx in range(5):
    row = 10 + idx
    rank = idx + 1
    amount_formula = f'IFERROR(LARGE(IF(\'Boletos a Pagar- AGOSTO\'!$A$12:$A$1000="✓",\'Boletos a Pagar- AGOSTO\'!$G$12:$G$1000),{rank}),"")'
    name_formula = f'IFERROR(INDEX(\'Boletos a Pagar- AGOSTO\'!$C$12:$C$1000,MATCH(AB{row},\'Boletos a Pagar- AGOSTO\'!$G$12:$G$1000,0)),"")'
    sheet = replace_formula_and_cache(sheet, f'AB{row}', amount_formula, top[idx][0] or '')
    sheet = replace_formula_and_cache(sheet, f'AC{row}', name_formula, top[idx][1])
    # Preserve the visible table formulas M=AC and N=AB, only refresh their cached display values.
    sheet = replace_formula_and_cache(sheet, f'M{row}', f'AC{row}', top[idx][1])
    sheet = replace_formula_and_cache(sheet, f'N{row}', f'AB{row}', top[idx][0] or 0)
files['xl/worksheets/sheet1.xml'] = sheet.encode('utf-8')

# Update only the existing Boletos Pagos chart's series references, keeping the chart object/style intact.
chart = files['xl/charts/chart3.xml'].decode('utf-8')
series = re.findall(r'<c:ser>.*?</c:ser>', chart)
assert len(series) >= 4
for idx in range(4):
    row = 10 + idx
    block = series[idx]
    block, n = re.subn(r'(<c:tx>.*?<c:f>)[^<]*(</c:f>)', rf'\1DASHBOARD!$M${row}\2', block, count=1, flags=re.S)
    assert n == 1
    block, n = re.subn(r'(<c:cat>.*?<c:f>)[^<]*(</c:f>)', rf'\1DASHBOARD!$M${row}\2', block, count=1, flags=re.S)
    assert n == 1
    block, n = re.subn(r'(<c:val>.*?<c:f>)[^<]*(</c:f>)', rf'\1DASHBOARD!$N${row}\2', block, count=1, flags=re.S)
    assert n == 1
    chart = chart.replace(series[idx], block, 1)
files['xl/charts/chart3.xml'] = chart.encode('utf-8')

with ZipFile(output, 'w', ZIP_DEFLATED) as zout:
    for name, data in files.items():
        zout.writestr(name, data)
print(output)
print('PAID_ROWS_FOUND', len(paid), 'CHART_SERIES_CONNECTED', 4, 'OTHER_CHART_OBJECTS_PRESERVED', True)
