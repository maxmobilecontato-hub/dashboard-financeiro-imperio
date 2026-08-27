from pathlib import Path
from zipfile import ZipFile
from openpyxl import load_workbook

output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_conectados.xlsx')
wb = load_workbook(output, data_only=False, read_only=False)
ws = wb['DASHBOARD']
assert len(ws._charts) == 4
for cell in ('E5', 'I5', 'M5', 'D16', 'D17', 'D18'):
    value = ws[cell].value
    assert isinstance(value, str) and value.startswith('=') and '#REF!' not in value, (cell, value)
for sheet, coords in {'Boletos a Pagar- AGOSTO':['A12','H12','J16','J40'], 'CONTROLE - fluxo de saida':['E18'], 'CONTROLE- Fluxo de Entrada':['E15']}.items():
    target = wb[sheet]
    for coord in coords:
        value = target[coord].value
        if isinstance(value, str): assert '#REF!' not in value, (sheet, coord, value)
# Inspect the XLSX package XML without iterating through the 16,384 formatted columns.
with ZipFile(output) as package:
    for name in package.namelist():
        if name.endswith('.xml'):
            raw = package.read(name)
            assert b'#REF!' not in raw, name
print('OK charts=4 operational_formulas=True no_broken_refs=True status_formula=True')
