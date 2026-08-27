from pathlib import Path
from openpyxl import load_workbook

path = Path('/tmp/recalc-connected/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_conectados.xlsx')
formula_wb = load_workbook(path, data_only=False)
value_wb = load_workbook(path, data_only=True)
wsf = formula_wb['DASHBOARD']
wsv = value_wb['DASHBOARD']
assert len(wsf._charts) == 4
assert isinstance(wsf['E5'].value, str) and '#REF!' not in wsf['E5'].value
assert isinstance(wsf['I5'].value, str) and '#REF!' not in wsf['I5'].value
assert isinstance(wsf['M5'].value, str) and '#REF!' not in wsf['M5'].value
assert isinstance(formula_wb['Boletos a Pagar- AGOSTO']['A12'].value, str) and formula_wb['Boletos a Pagar- AGOSTO']['A12'].value.startswith('=')
print('charts=', len(wsf._charts), 'cached_totals=', wsv['E5'].value, wsv['I5'].value, wsv['M5'].value, 'status_formula=True')
