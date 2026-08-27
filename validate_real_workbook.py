from openpyxl import load_workbook
from pathlib import Path
p=Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
wb=load_workbook(p,data_only=True)

def nums(ws, col, start):
    values=[]
    for r in range(start,ws.max_row+1):
        v=ws[f'{col}{r}'].value
        if isinstance(v,(int,float)) and v>0: values.append(float(v))
    return values
entries=nums(wb['CONTROLE- Fluxo de Entrada'],'E',15)
expenses=nums(wb['CONTROLE - fluxo de saida'],'E',18)
bills=nums(wb['Boletos a Pagar- AGOSTO'],'G',12)
receivables=nums(wb['1- PAGAMENTOS A RECEBER '],'E',9)
print('entries_count',len(entries),'entries_total',sum(entries))
print('expenses_count',len(expenses),'expenses_total',sum(expenses))
print('bills_count',len(bills),'bills_total',sum(bills))
print('receivables_count',len(receivables),'receivables_total',sum(receivables))
print('balance',sum(entries)-sum(expenses)-sum(bills))
print('receivables_value_column','E')
