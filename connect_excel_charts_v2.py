from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_conectados.xlsx')
wb = load_workbook(source)
ws = wb['DASHBOARD']
bills = wb['Boletos a Pagar- AGOSTO']
ws._charts = []

# Headline totals linked directly to the operational ledgers.
ws['E5'] = "=SUM('CONTROLE - fluxo de saida'!$E$18:$E$1001)"
ws['I5'] = "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
ws['M5'] = "=SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)"

# Fix legacy helper formulas that contained broken #REF references, without touching visible status formulas.
for row in range(12, 1001):
    cell = bills[f'J{row}']
    if isinstance(cell.value, str) and '#REF!' in cell.value:
        previous = row - 1
        cell.value = f'=IF(ISNUMBER(SEARCH("PAGO",A{row}&" "&H{row}&" "&I{row})),COUNTIF($J$11:J{previous},">0")+1,0)'
for cf in bills.conditional_formatting:
    for rule in bills.conditional_formatting[cf]:
        if rule.formula:
            rule.formula = [formula.replace('#REF!', '$H12') for formula in rule.formula]
            if str(cf) == 'A39:I39' and rule.formula == ['$H12="VENCIDO"']:
                rule.formula = ['AND($F39<>"",$F39<TODAY(),$I39="....")']

# Payment-method summary remains directly linked to the exit ledger.
ws['D16'] = "=SUMIF('CONTROLE - fluxo de saida'!$D$15:$D$1001,C16,'CONTROLE - fluxo de saida'!$E$15:$E$1001)"
ws['D17'] = "=SUMIF('CONTROLE - fluxo de saida'!$D$15:$D$1001,C17,'CONTROLE - fluxo de saida'!$E$15:$E$1001)"
ws['D18'] = "=SUMIF('CONTROLE - fluxo de saida'!$D$15:$D$1001,\"*DEBITO*\",'CONTROLE - fluxo de saida'!$E$15:$E$1001)"

# Dynamic top-five paid-bill helper, replacing the previous #REF formulas.
for row in range(10, 15):
    ws[f'N{row}'] = f'=IFERROR(AGGREGATE(14,6,\'Boletos a Pagar- AGOSTO\'!$G$12:$G$1000/(\'Boletos a Pagar- AGOSTO\'!$A$12:$A$1000="✓"),ROWS($N$10:N{row})),"")'
    ws[f'M{row}'] = f'=IFERROR(INDEX(\'Boletos a Pagar- AGOSTO\'!$C$12:$C$1000,AGGREGATE(15,6,(ROW(\'Boletos a Pagar- AGOSTO\'!$C$12:$C$1000)-ROW(\'Boletos a Pagar- AGOSTO\'!$C$12)+1)/((\'Boletos a Pagar- AGOSTO\'!$A$12:$A$1000="✓")*(\'Boletos a Pagar- AGOSTO\'!$G$12:$G$1000=N{row})),1)),"")'
    ws[f'AB{row}'] = f'=N{row}'
    ws[f'AC{row}'] = f'=M{row}'

# Helper table for the overview chart.
for row, label, formula in [(17, 'Indicador', 'Valor'), (18, 'Total de entradas', '=I5'), (19, 'Total de saídas', '=E5'), (20, 'Boletos pagos', '=M5')]:
    ws.cell(row=row, column=24).value = label
    ws.cell(row=row, column=25).value = formula
    for col in (24, 25):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(color='FFFFFF', bold=row == 17)
        cell.fill = PatternFill(fill_type='solid', fgColor='11182E')
        cell.alignment = Alignment(horizontal='left')

# Dedicated helper table for the payment composition chart.
for row, label, formula in [(22, 'Forma de pagamento', 'Total'), (23, '=C16', '=D16'), (24, '=C17', '=D17'), (25, '=C18', '=D18')]:
    ws.cell(row=row, column=24).value = label
    ws.cell(row=row, column=25).value = formula

# Overview chart: direct linked totals.
overview = BarChart()
overview.type = 'col'
overview.style = 10
overview.title = 'Resumo financeiro atualizado'
overview.y_axis.title = 'Valor (R$)'
overview.height = 8.0
overview.width = 13.0
overview.add_data(Reference(ws, min_col=25, min_row=17, max_row=20), titles_from_data=True)
overview.set_categories(Reference(ws, min_col=24, min_row=18, max_row=20))
overview.legend = None
overview.dataLabels = DataLabelList()
overview.dataLabels.showVal = True
if overview.series: overview.series[0].graphicalProperties.solidFill = '35B8C9'
ws.add_chart(overview, 'C17')

# Composition chart: direct linked payment-method totals.
payment = PieChart()
payment.title = 'Despesas por forma de pagamento'
payment.style = 10
payment.height = 8.0
payment.width = 13.0
payment.add_data(Reference(ws, min_col=25, min_row=22, max_row=25), titles_from_data=True)
payment.set_categories(Reference(ws, min_col=24, min_row=23, max_row=25))
payment.dataLabels = DataLabelList()
payment.dataLabels.showPercent = True
payment.dataLabels.showLeaderLines = True
ws.add_chart(payment, 'M17')

# Top five exits: dashboard cells are driven by the existing dynamic X/Y helpers.
expenses = BarChart()
expenses.type = 'bar'
expenses.style = 10
expenses.title = 'Top 5 — principais saídas'
expenses.x_axis.title = 'Valor (R$)'
expenses.height = 8.0
expenses.width = 13.0
expenses.add_data(Reference(ws, min_col=4, min_row=9, max_row=14), titles_from_data=True)
expenses.set_categories(Reference(ws, min_col=3, min_row=10, max_row=14))
expenses.legend = None
expenses.dataLabels = DataLabelList()
expenses.dataLabels.showVal = True
if expenses.series: expenses.series[0].graphicalProperties.solidFill = 'C451A0'
ws.add_chart(expenses, 'C35')

# Top five entries: dashboard cells are driven by the existing dynamic Z/AA helpers.
entries = BarChart()
entries.type = 'bar'
entries.style = 10
entries.title = 'Top 5 — principais entradas'
entries.x_axis.title = 'Valor (R$)'
entries.height = 8.0
entries.width = 13.0
entries.add_data(Reference(ws, min_col=10, min_row=9, max_row=14), titles_from_data=True)
entries.set_categories(Reference(ws, min_col=9, min_row=10, max_row=14))
entries.legend = None
entries.dataLabels = DataLabelList()
entries.dataLabels.showVal = True
if entries.series: entries.series[0].graphicalProperties.solidFill = '35B8C9'
ws.add_chart(entries, 'M35')

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
except AttributeError:
    pass

wb.save(output)
print(output)
print('DASHBOARD_CHARTS', len(ws._charts))
print('DYNAMIC_LINKS', ws['E5'].value, ws['I5'].value, ws['M5'].value)
