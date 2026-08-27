from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment

source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
output = Path('/home/ubuntu/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(source)
ws = wb['DASHBOARD']
ws._charts = []

# Helper table stays in the existing dashboard support area and feeds the overview chart.
helper = [('Indicador', 'Valor'), ('Total de entradas', '=J5'), ('Total de saídas', '=G5'), ('Boletos pagos', '=N5')]
for row, values in enumerate(helper, start=17):
    ws.cell(row=row, column=25).value = values[0]
    ws.cell(row=row, column=26).value = values[1]
    for col in (25, 26):
        ws.cell(row=row, column=col).font = Font(color='FFFFFF', bold=row == 17)
        ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='11182E')
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='left')

# Overall comparison: the most important daily reading.
overview = BarChart()
overview.type = 'col'
overview.style = 10
overview.title = 'Resumo financeiro'
overview.y_axis.title = 'Valor (R$)'
overview.height = 8.0
overview.width = 13.5
overview.add_data(Reference(ws, min_col=26, min_row=17, max_row=20), titles_from_data=True)
overview.set_categories(Reference(ws, min_col=25, min_row=18, max_row=20))
overview.legend = None
overview.dataLabels = DataLabelList()
overview.dataLabels.showVal = True
if overview.series:
    overview.series[0].graphicalProperties.solidFill = '35B8C9'
ws.add_chart(overview, 'C17')

# Composition: better than a long table for payment-method mix.
payment = PieChart()
payment.title = 'Despesas por forma de pagamento'
payment.style = 10
payment.height = 8.0
payment.width = 13.5
payment.add_data(Reference(ws, min_col=4, min_row=8, max_row=11), titles_from_data=True)
payment.set_categories(Reference(ws, min_col=3, min_row=9, max_row=11))
payment.dataLabels = DataLabelList()
payment.dataLabels.showPercent = True
payment.dataLabels.showLeaderLines = True
ws.add_chart(payment, 'M17')

# Ranking: horizontal bars make long names readable.
expenses = BarChart()
expenses.type = 'bar'
expenses.style = 10
expenses.title = 'Top 5 — principais saídas'
expenses.x_axis.title = 'Valor (R$)'
expenses.height = 8.0
expenses.width = 13.5
expenses.add_data(Reference(ws, min_col=7, min_row=9, max_row=14), titles_from_data=True)
expenses.set_categories(Reference(ws, min_col=6, min_row=10, max_row=14))
expenses.legend = None
expenses.dataLabels = DataLabelList()
expenses.dataLabels.showVal = True
if expenses.series:
    expenses.series[0].graphicalProperties.solidFill = 'C451A0'
ws.add_chart(expenses, 'C35')

entries = BarChart()
entries.type = 'bar'
entries.style = 10
entries.title = 'Top 5 — principais entradas'
entries.x_axis.title = 'Valor (R$)'
entries.height = 8.0
entries.width = 13.5
entries.add_data(Reference(ws, min_col=11, min_row=8, max_row=14), titles_from_data=True)
entries.set_categories(Reference(ws, min_col=10, min_row=10, max_row=14))
entries.legend = None
entries.dataLabels = DataLabelList()
entries.dataLabels.showVal = True
if entries.series:
    entries.series[0].graphicalProperties.solidFill = '35B8C9'
ws.add_chart(entries, 'M35')

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
except AttributeError:
    pass

wb.save(output)
print(output)
print('DASHBOARD_CHARTS', len(ws._charts))
print('ANCHORS', ['C17', 'M17', 'C35', 'M35'])
