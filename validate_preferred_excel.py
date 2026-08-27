from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
output = Path('/home/ubuntu/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
original = load_workbook(source, data_only=False)
restored = load_workbook(output, data_only=False)
assert restored.sheetnames == original.sheetnames
ws = restored['DASHBOARD']
assert len(ws._charts) == 4
anchors = []
for chart in ws._charts:
    anchor = chart.anchor
    anchors.append(anchor if isinstance(anchor, str) else ws.cell(row=anchor._from.row + 1, column=anchor._from.col + 1).coordinate)
assert anchors == ['C17', 'M17', 'C35', 'M35'], anchors
for sheet, coords in {'DASHBOARD': ['D9', 'D10', 'D11', 'G5', 'J5', 'N5'], 'Boletos a Pagar- AGOSTO': ['A12', 'H12']}.items():
    for coord in coords:
        assert restored[sheet][coord].value == original[sheet][coord].value, (sheet, coord)
assert len(restored['Boletos a Pagar- AGOSTO'].conditional_formatting) == len(original['Boletos a Pagar- AGOSTO'].conditional_formatting)
assert isinstance(restored['Boletos a Pagar- AGOSTO']['A12'].value, str) and restored['Boletos a Pagar- AGOSTO']['A12'].value.startswith('=')
assert isinstance(restored['Boletos a Pagar- AGOSTO']['H12'].value, str) and restored['Boletos a Pagar- AGOSTO']['H12'].value.startswith('=')
print('OK charts=4 anchors=C17,M17,C35,M35 formulas_preserved=True status_automation_preserved=True conditional_formatting_preserved=True')
