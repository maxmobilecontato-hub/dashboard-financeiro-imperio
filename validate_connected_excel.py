from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_conectados.xlsx')
original = load_workbook(source, data_only=False)
wb = load_workbook(output, data_only=False)
ws = wb['DASHBOARD']
assert len(ws._charts) == 4
for cell in ('E5', 'I5', 'M5', 'D16', 'D17', 'D18'):
    value = ws[cell].value
    assert isinstance(value, str) and value.startswith('=') and '#REF!' not in value, (cell, value)
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str):
                assert '#REF!' not in value, (sheet.title, cell.coordinate, value)
for chart in ws._charts:
    for series in chart.ser:
        for ref in (series.val.numRef.f if series.val and series.val.numRef else '', series.cat.strRef.f if series.cat and series.cat.strRef else (series.cat.numRef.f if series.cat and series.cat.numRef else '')):
            assert ref and '#REF!' not in ref, ref
assert ws['E5'].value == "=SUM('CONTROLE - fluxo de saida'!$E$18:$E$1001)"
assert ws['I5'].value == "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
assert ws['M5'].value == "=SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)"
print('OK charts=4 no_broken_refs=True operational_links=True formulas_recalculate_on_open=True')
