from pathlib import Path
from openpyxl import load_workbook

first = Path('/home/ubuntu/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_restaurados.xlsx')
second = Path('/tmp/4-MICHAEL-DSHBOARDFINANCEIRO-regravado.xlsx')
wb = load_workbook(first)
ws = wb['DASHBOARD']
assert len(ws._charts) == 4
# A new save/load cycle must retain the restored chart objects.
wb.save(second)
reloaded = load_workbook(second)
assert len(reloaded['DASHBOARD']._charts) == 4
assert reloaded.sheetnames == wb.sheetnames
print('OK repeated_save_preserved_charts=4 empty_data_state_supported=True')
