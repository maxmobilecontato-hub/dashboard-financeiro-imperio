from openpyxl import load_workbook
from pathlib import Path
source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
wb = load_workbook(source, data_only=False)
ws = wb['DASHBOARD']
for row in range(1, ws.max_row + 1):
    cells = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value is not None:
            cells.append(f'{ws.cell(row,col).coordinate}={value}')
    if cells:
        print(' | '.join(cells))
