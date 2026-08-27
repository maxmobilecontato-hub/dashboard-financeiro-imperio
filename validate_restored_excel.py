from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
output = Path('/home/ubuntu/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_restaurados.xlsx')
original = load_workbook(source, data_only=False)
restored = load_workbook(output, data_only=False)
assert restored.sheetnames == original.sheetnames
assert len(restored['DASHBOARD']._charts) == 4
for name in original.sheetnames:
    before, after = original[name], restored[name]
    print('dimensions', name, (before.max_row, before.max_column), (after.max_row, after.max_column))
    assert after.max_row >= before.max_row and after.max_column >= before.max_column
    assert len(before.conditional_formatting) == len(after.conditional_formatting)
for coord in ('D9', 'D10', 'D11', 'J5', 'N5'):
    assert restored['DASHBOARD'][coord].value is not None
print('OK charts=4 sheets=', len(restored.sheetnames), 'conditional_formatting_preserved=True')
print('chart_titles=', [chart.title.tx.rich.p[0].r[0].t for chart in restored['DASHBOARD']._charts if chart.title and chart.title.tx and chart.title.tx.rich])
