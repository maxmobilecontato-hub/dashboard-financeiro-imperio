from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
output = Path('/home/ubuntu/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_restaurados.xlsx')
original = load_workbook(source, data_only=False)
restored = load_workbook(output, data_only=False)

assert restored.sheetnames == original.sheetnames
assert len(restored['DASHBOARD']._charts) == 4

# Key dashboard formulas remain unchanged after restoring chart objects.
for sheet, coords in {
    'DASHBOARD': ['D9', 'D10', 'D11', 'J5', 'N5', 'F10', 'G10', 'J10', 'K10', 'N10', 'O10'],
    'Boletos a Pagar- AGOSTO': ['A12', 'H12'],
}.items():
    for coord in coords:
        assert restored[sheet][coord].value == original[sheet][coord].value, (sheet, coord)

# Status automation remains formula-driven and conditional formatting is preserved.
bills_before = original['Boletos a Pagar- AGOSTO']
bills_after = restored['Boletos a Pagar- AGOSTO']
assert isinstance(bills_after['A12'].value, str) and bills_after['A12'].value.startswith('=')
assert isinstance(bills_after['H12'].value, str) and bills_after['H12'].value.startswith('=')
assert len(bills_after.conditional_formatting) == len(bills_before.conditional_formatting)
assert len(bills_after.merged_cells.ranges) == len(bills_before.merged_cells.ranges)
print('OK formulas_preserved=True status_formulas_preserved=True conditional_formatting_preserved=True charts=4')
