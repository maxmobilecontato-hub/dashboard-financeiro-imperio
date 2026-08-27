from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
wb = load_workbook(source, data_only=False)
print('SHEETS', wb.sheetnames)
for ws in wb.worksheets:
    print('SHEET', ws.title, 'DIM', ws.max_row, ws.max_column, 'CHARTS', len(ws._charts), 'MERGES', len(ws.merged_cells.ranges))
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        values = [str(v)[:80] if v is not None else '' for v in row[:min(ws.max_column, 12)]]
        if any(values): print('ROW', values)
    print()
