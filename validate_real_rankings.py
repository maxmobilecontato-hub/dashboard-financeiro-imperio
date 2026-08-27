from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
wb = load_workbook(path, data_only=True, read_only=True)

def top_from_sheet(title, header_row, name_col, amount_col):
    rows = list(wb[title].iter_rows(min_row=header_row + 1, values_only=True))
    grouped = defaultdict(float)
    for row in rows:
        name = str(row[name_col] or '').strip()
        amount = row[amount_col]
        if name and isinstance(amount, (int, float)) and amount > 0:
            grouped[name] += float(amount)
    return sorted(grouped.items(), key=lambda item: item[1], reverse=True)[:5]

entries = top_from_sheet('CONTROLE- Fluxo de Entrada', 14, 1, 4)
expenses = top_from_sheet('CONTROLE - fluxo de saida', 17, 1, 4)
print('entries_top5=', entries)
print('expenses_top5=', expenses)
print('entries_total=', sum(value for _, value in entries))
print('expenses_total_top5=', sum(value for _, value in expenses))
