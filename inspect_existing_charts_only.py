from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(path, data_only=False)
print('SHEETS', wb.sheetnames)
for ws in wb.worksheets:
    if ws._charts:
        print('SHEET', ws.title, 'CHARTS', len(ws._charts))
        for i, chart in enumerate(ws._charts, 1):
            anchor = chart.anchor
            title = ''
            try:
                title = chart.title.tx.rich.p[0].r[0].t
            except Exception:
                pass
            print('CHART', i, 'title=', title, 'anchor=', getattr(anchor, '_from', None), 'editAs=', getattr(anchor, 'editAs', None))
            for s in chart.ser:
                values = getattr(getattr(s, 'val', None), 'numRef', None)
                cats = getattr(getattr(s, 'cat', None), 'strRef', None) or getattr(getattr(s, 'cat', None), 'numRef', None)
                print('  values=', getattr(values, 'f', None), 'categories=', getattr(cats, 'f', None))
ws = wb['DASHBOARD']
for cell in ('E5','I5','M5','Y18','Y19','Y20','D16','D17','D18'):
    print('CELL', cell, repr(ws[cell].value))
