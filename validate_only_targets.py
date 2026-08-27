from pathlib import Path
from zipfile import ZipFile
from openpyxl import load_workbook

original = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
final = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_originais_conectados.xlsx')
wo = load_workbook(original, data_only=False)
wf = load_workbook(final, data_only=False)
so, sf = wo['DASHBOARD'], wf['DASHBOARD']
assert len(so._charts) == len(sf._charts) == 2
for co, cf in zip(so._charts, sf._charts):
    assert type(co) is type(cf)
    assert co.anchor._from.col == cf.anchor._from.col and co.anchor._from.row == cf.anchor._from.row
    assert co.anchor._from.colOff == cf.anchor._from.colOff and co.anchor._from.rowOff == cf.anchor._from.rowOff
    assert co.title.tx.rich.p[0].r[0].t == cf.title.tx.rich.p[0].r[0].t
    for so_ser, sf_ser in zip(co.ser, cf.ser):
        ov = getattr(getattr(so_ser, 'val', None), 'numRef', None)
        fv = getattr(getattr(sf_ser, 'val', None), 'numRef', None)
        oc = getattr(getattr(so_ser, 'cat', None), 'strRef', None) or getattr(getattr(so_ser, 'cat', None), 'numRef', None)
        fc = getattr(getattr(sf_ser, 'cat', None), 'strRef', None) or getattr(getattr(sf_ser, 'cat', None), 'numRef', None)
        assert getattr(ov, 'f', None) == getattr(fv, 'f', None)
        assert getattr(oc, 'f', None) == getattr(fc, 'f', None)
assert sf['I5'].value == "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
assert sf['M5'].value == "=SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)"
def ref_files(path):
    with ZipFile(path) as z:
        return {name: z.read(name).count(b'#REF!') for name in z.namelist() if name.endswith('.xml')}
original_refs = ref_files(original)
final_refs = ref_files(final)
assert {k: v for k, v in original_refs.items() if v} == {k: v for k, v in final_refs.items() if v}
print('OK charts_unchanged=True chart_count=2 target_formulas_connected=True existing_refs_preserved=True')
