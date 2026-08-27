from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(path, data_only=False)
ws = wb['DASHBOARD']
for r in range(1, 25):
    values = [ws.cell(r, c).value for c in range(24, 31)]
    if any(v is not None for v in values): print(r, values)
