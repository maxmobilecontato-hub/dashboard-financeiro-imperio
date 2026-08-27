from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_originais_conectados.xlsx')
wb = load_workbook(source)
ws = wb['DASHBOARD']
chart_count_before = len(ws._charts)

# Only connect the two requested dashboard indicators to operational sheets.
ws['I5'] = "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
ws['M5'] = "=SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)"
# Preserve the existing helper links used by the dashboard summary.
ws['Y18'] = '=I5'
ws['Y20'] = '=M5'

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
except AttributeError:
    pass

wb.save(output)
reloaded = load_workbook(output, data_only=False)
assert len(reloaded['DASHBOARD']._charts) == chart_count_before
assert reloaded['DASHBOARD']['I5'].value == "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
assert reloaded['DASHBOARD']['M5'].value == "=SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)"
print(output)
print('CHARTS_PRESERVED', chart_count_before)
print('TARGETS_CONNECTED', 'I5', 'M5')
