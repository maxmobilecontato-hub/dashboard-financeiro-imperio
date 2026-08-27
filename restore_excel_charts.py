from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment

source = Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
output = Path('/home/ubuntu/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_restaurados.xlsx')
wb = load_workbook(source)
ws = wb['DASHBOARD']

# Remove only chart objects from the dashboard so rerunning the script is safe.
ws._charts = []

# Keep a compact, neutral visual language compatible with the existing dashboard.
for cell in ('Q1', 'Q17', 'Q33', 'Q49'):
    ws[cell].fill = PatternFill(fill_type='solid', fgColor='11182E')
    ws[cell].font = Font(color='FFFFFF', bold=True)
    ws[cell].alignment = Alignment(horizontal='left')

# 1) Expenses by payment method: C9:D11.
payment_chart = PieChart()
payment_chart.title = 'Despesas por forma de pagamento'
payment_chart.style = 10
payment_chart.height = 7.0
payment_chart.width = 12.0
payment_chart.add_data(Reference(ws, min_col=4, min_row=8, max_row=11), titles_from_data=True)
payment_chart.set_categories(Reference(ws, min_col=3, min_row=9, max_row=11))
payment_chart.dataLabels = DataLabelList()
payment_chart.dataLabels.showPercent = True
ws.add_chart(payment_chart, 'Q2')

# 2) Main outgoing launches: F9:G14.
expense_chart = BarChart()
expense_chart.type = 'bar'
expense_chart.style = 10
expense_chart.title = 'Principais lançamentos de saída'
expense_chart.y_axis.title = 'Lançamento'
expense_chart.x_axis.title = 'Valor'
expense_chart.height = 8.0
expense_chart.width = 14.0
expense_chart.add_data(Reference(ws, min_col=7, min_row=9, max_row=14), titles_from_data=True)
expense_chart.set_categories(Reference(ws, min_col=6, min_row=10, max_row=14))
expense_chart.legend = None
ws.add_chart(expense_chart, 'Q18')

# 3) Main incoming launches: J8:K14.
entry_chart = BarChart()
entry_chart.type = 'col'
entry_chart.style = 10
entry_chart.title = 'Principais lançamentos de entrada'
entry_chart.y_axis.title = 'Valor'
entry_chart.x_axis.title = 'Lançamento'
entry_chart.height = 8.0
entry_chart.width = 14.0
entry_chart.add_data(Reference(ws, min_col=11, min_row=8, max_row=14), titles_from_data=True)
entry_chart.set_categories(Reference(ws, min_col=10, min_row=10, max_row=14))
entry_chart.legend = None
ws.add_chart(entry_chart, 'Q34')

# 4) Paid bills: N9:O14.
paid_chart = BarChart()
paid_chart.type = 'bar'
paid_chart.style = 10
paid_chart.title = 'Boletos pagos'
paid_chart.y_axis.title = 'Lançamento'
paid_chart.x_axis.title = 'Valor'
paid_chart.height = 8.0
paid_chart.width = 14.0
paid_chart.add_data(Reference(ws, min_col=15, min_row=9, max_row=14), titles_from_data=True)
paid_chart.set_categories(Reference(ws, min_col=14, min_row=10, max_row=14))
paid_chart.legend = None
ws.add_chart(paid_chart, 'Q50')

# Ask Excel to recalculate the existing formulas when the restored workbook opens.
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
except AttributeError:
    pass

wb.save(output)
print(output)
print('DASHBOARD_CHARTS', len(ws._charts))
print('SHEETS_UNCHANGED', wb.sheetnames)
