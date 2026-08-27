from pathlib import Path
from shutil import copy2
from zipfile import ZipFile
from openpyxl import load_workbook
import re

source = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_boleto_grafico_conectado.xlsx')
temp = Path('/tmp/boleto_payment_update_test.xlsx')
copy2(source, temp)
wb = load_workbook(temp, data_only=False)
ws = wb['Boletos a Pagar- AGOSTO']
# Simulate adding a payment proof to a previously unpaid row.
ws['I17'] = 'COMPROVANTE TESTE'
wb.save(temp)
reloaded = load_workbook(temp, data_only=False)
bills = reloaded['Boletos a Pagar- AGOSTO']
assert bills['I17'].value == 'COMPROVANTE TESTE'
assert '$I17' in bills['A17'].value and 'PAGO' in bills['H17'].value
with ZipFile(temp) as z:
    sheet = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
    chart = z.read('xl/charts/chart3.xml').decode('utf-8')
    assert 'Boletos a Pagar- AGOSTO' in sheet
    assert 'Boletos a Pagar- AGOSTO' in sheet and '$A$12:$A$1000' in sheet and '$G$12:$G$1000' in sheet
    for row in range(10, 14):
        assert f'DASHBOARD!$M${row}' in chart
        assert f'DASHBOARD!$N${row}' in chart
print('OK simulated_payment_proof=True status_formula_preserved=True boleto_ranking_formula_preserved=True chart_source_preserved=True')
temp.unlink(missing_ok=True)
