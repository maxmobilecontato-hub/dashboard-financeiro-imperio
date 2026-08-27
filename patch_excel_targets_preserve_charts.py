from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl import load_workbook
import re

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_preservados_conectados.xlsx')

# Compute current cache values from the real source workbook.
values_wb = load_workbook(source, data_only=True, read_only=True)
entry_ws = values_wb['CONTROLE- Fluxo de Entrada']
bills_ws = values_wb['Boletos a Pagar- AGOSTO']
total_entries = sum((cell.value or 0) for row in entry_ws.iter_rows(min_row=15, max_row=1000, min_col=5, max_col=5) for cell in row if isinstance(cell.value, (int, float)))
total_paid_bills = sum((bills_ws.cell(row, 7).value or 0) for row in range(12, 1001) if bills_ws.cell(row, 1).value == '✓' and isinstance(bills_ws.cell(row, 7).value, (int, float)))
values_wb.close()

with ZipFile(source, 'r') as zin:
    original_files = {info.filename: zin.read(info.filename) for info in zin.infolist()}
xml = original_files['xl/worksheets/sheet1.xml'].decode('utf-8')

def replace_cell(xml_text, cell, formula, cached):
    pattern = rf'(<c r="{cell}"[^>]*>)(.*?)(</c>)'
    replacement = rf'\1<f>{formula}</f><v>{cached}</v>\3'
    new_xml, count = re.subn(pattern, replacement, xml_text, count=1)
    assert count == 1, cell
    return new_xml

xml = replace_cell(xml, 'I5', "SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)", total_entries)
xml = replace_cell(xml, 'M5', "SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)", total_paid_bills)
original_files['xl/worksheets/sheet1.xml'] = xml.encode('utf-8')

with ZipFile(output, 'w', ZIP_DEFLATED) as zout:
    for name, data in original_files.items():
        zout.writestr(name, data)

with ZipFile(output) as check:
    assert check.read('xl/charts/chart1.xml') == original_files.get('xl/charts/chart1.xml')
    assert check.read('xl/charts/chart2.xml') == original_files.get('xl/charts/chart2.xml')
    assert check.read('xl/charts/chartEx1.xml') == original_files.get('xl/charts/chartEx1.xml')
print(output)
print('TOTAL_ENTRIES_CACHE', total_entries)
print('PAID_BILLS_CACHE', total_paid_bills)
print('CHART_XML_PRESERVED', True)
