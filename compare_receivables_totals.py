from openpyxl import load_workbook
from pathlib import Path
p=Path('/home/ubuntu/upload/4-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_status_cores_funcionais.xlsx')
wb=load_workbook(p,data_only=True)
ws=wb['1- PAGAMENTOS A RECEBER ']
valid=[]
for r in range(9, ws.max_row+1):
    name=ws.cell(r,3).value
    amount=ws.cell(r,5).value
    if isinstance(name,str) and name.strip() and isinstance(amount,(int,float)) and amount>0:
        valid.append((name.strip(),float(amount)))
direct_total=sum(amount for _,amount in valid)
parser_total=1138.71
print({'direct_valid_count':len(valid),'direct_valid_total':direct_total,'parser_total':parser_total,'equal':round(direct_total,2)==round(parser_total,2),'valid_records':valid})
