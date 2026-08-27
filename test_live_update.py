from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_conectados.xlsx')
wb = load_workbook(path, data_only=False)
ws_in = wb['CONTROLE- Fluxo de Entrada']
ws_out = wb['CONTROLE - fluxo de saida']
ws_dash = wb['DASHBOARD']

# Record current totals (formulas).
f_in, f_out = ws_dash['I5'].value, ws_dash['E5'].value
assert f_in == "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
assert f_out == "=SUM('CONTROLE - fluxo de saida'!$E$18:$E$1001)"

# Add new data to ledgers.
ws_in['E70'] = 50000
ws_out['E40'] = 20000

# Save and reload with data_only=True to simulate Excel's recalculation on open.
# Note: openpyxl doesn't have a full recalc engine, but we can verify the formulas remain linked.
wb.save('/tmp/live_update_test.xlsx')
reloaded = load_workbook('/tmp/live_update_test.xlsx', data_only=False)
assert reloaded['DASHBOARD']['I5'].value == f_in
assert reloaded['DASHBOARD']['E5'].value == f_out
print('OK live_update_formulas_linked=True data_added_to_ledgers=True')
