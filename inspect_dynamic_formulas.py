from pathlib import Path
from openpyxl import load_workbook
source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(source, data_only=False)
ws = wb['DASHBOARD']
for coord in ['X10','Y10','Z10','AA10','AB10','AC10','X11','Y11','Z11','AA11','AB11','AC11']:
    value = ws[coord].value
    print(coord, getattr(value, 'text', value))
for sheet, coords in {'CONTROLE - fluxo de saida':['D15','E15'], 'CONTROLE- Fluxo de Entrada':['D15','E15'], 'Boletos a Pagar- AGOSTO':['A12','H12']}.items():
    ws2 = wb[sheet]
    print(sheet)
    for coord in coords: print(coord, ws2[coord].value)
