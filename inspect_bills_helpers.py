from pathlib import Path
from openpyxl import load_workbook
source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(source, data_only=False)
ws = wb['Boletos a Pagar- AGOSTO']
for row in range(1, 48):
    values = []
    for col in range(1, 13):
        value = ws.cell(row, col).value
        if value is not None:
            values.append(f'{ws.cell(row,col).coordinate}={getattr(value, "text", value)}')
    if values: print(' | '.join(values))
