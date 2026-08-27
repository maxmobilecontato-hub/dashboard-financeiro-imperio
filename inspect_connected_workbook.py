from pathlib import Path
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
wb = load_workbook(source, data_only=False)
ws = wb['DASHBOARD']
print('SHEETS', wb.sheetnames)
print('CHARTS', len(ws._charts))
for i, chart in enumerate(ws._charts, 1):
    print('CHART', i, 'anchor', chart.anchor, 'title', chart.title.tx.rich.p[0].r[0].t if chart.title and chart.title.tx and chart.title.tx.rich else '')
    for series in chart.ser:
        print('  SERIES', series.tx, 'values', series.val.numRef.f if series.val and series.val.numRef else '', 'categories', series.cat.strRef.f if series.cat and series.cat.strRef else (series.cat.numRef.f if series.cat and series.cat.numRef else ''))
for coord in ('G5', 'J5', 'N5', 'C9', 'D9', 'D10', 'D11', 'F10', 'G10', 'J10', 'K10'):
    print('CELL', coord, ws[coord].value)
