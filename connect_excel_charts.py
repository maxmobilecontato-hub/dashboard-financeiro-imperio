from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment

source = Path('/home/ubuntu/upload/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_aprimorados.xlsx')
output = Path('/home/ubuntu/5-MICHAEL-DSHBOARDFINANCEIRO-25-08-2026_graficos_conectados.xlsx')
wb = load_workbook(source)
ws = wb['DASHBOARD']
ws._charts = []

# Normalize the three headline totals so they always read from the operational sheets.
ws['E5'] = "=SUM('CONTROLE - fluxo de saida'!$E$18:$E$1001)"
ws['I5'] = "=SUM('CONTROLE- Fluxo de Entrada'!$E$15:$E$1000)"
ws['M5'] = "=SUMIF('Boletos a Pagar- AGOSTO'!$A$12:$A$1000,\"✓\",'Boletos a Pagar- AGOSTO'!$G$12:$G$1000)"

# A small source table for the overview chart. Its values are formulas linked to the headline totals.
for row, label, formula in [(17, 'Indicador', 'Valor'), (18, 'Total de entradas', '=I5'), (19, 'Total de saídas', '=E5'), (20, 'Boletos pagos', '=M5')]:
    ws.cell(row=row, column=24).value = label
    ws.cell(row=row, column=25).value = formula
    for col in (24, 25):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(color='FFFFFF', bold=row == 17)
        cell.fill = PatternFill(fill_type='solid', fgColor='11182E')
        cell.alignment = Alignment(horizontal='left')

# Ensure the existing payment-method summary remains linked to the exit ledger.
ws['D16'] = "=SUMIF('CONTROLE - fluxo de saida'!$D$15:$D$1001,C16,'CONTROLE - fluxo de saida'!$E$15:$E$1001)"
ws['D17'] = "=SUMIF('CONTROLE - fluxo de saida'!$D$15:$D$1001,C17,'CONTROLE - fluxo de saida'!$E$15:$E$1001)"
ws['D18'] = "=SUMIF('CONTROLE - fluxo de saida'!$D$15:$D$1001,\"*DEBITO*\",'CONTROLE - fluxo de saida'!$E$15:$E$1001)"

# 1. Overview: entries versus exits versus paid bills.
overview = BarChart()
overview.type = 'col'
overview.style = 10
overview.title = 'Resumo financeiro atualizado'
overview.y_axis.title = 'Valor (R$)'
overview.height = 8.0
overview.width = 13.0
overview.add_data(Reference(ws, min_col=25, min_row=17, max_row=20), titles_from_data=True)
overview.set_categories(Reference(ws, min_col=24, min_row=18, max_row=20))
overview.legend = None
overview.dataLabels = DataLabelList()
overview.dataLabels.showVal = True
if overview.series:
    overview.series[0].graphicalProperties.solidFill = '35B8C9'
ws.add_chart(overview, 'C17')

# 2. Composition of exits by payment method.
payment = PieChart()
payment.title = 'Despesas por forma de pagamento'
payment.style = 10
payment.height = 8.0
payment.width = 13.0
payment.add_data(Reference(ws, min_col=4, min_row=8, max_row=18), titles_from_data=True)
payment.set_categories(Reference(ws, min_col=3, min_row=16, max_row=18))
payment.dataLabels = DataLabelList()
payment.dataLabels.showPercent = True
payment.dataLabels.showLeaderLines = True
ws.add_chart(payment, 'M17')

# 3. Top outgoing launches. C10:D14 are fed by the dynamic helper formulas X/Y.
expenses = BarChart()
expenses.type = 'bar'
expenses.style = 10
expenses.title = 'Top 5 — principais saídas'
expenses.x_axis.title = 'Valor (R$)'
expenses.height = 8.0
expenses.width = 13.0
expenses.add_data(Reference(ws, min_col=4, min_row=9, max_row=14), titles_from_data=True)
expenses.set_categories(Reference(ws, min_col=3, min_row=10, max_row=14))
expenses.legend = None
expenses.dataLabels = DataLabelList()
expenses.dataLabels.showVal = True
if expenses.series:
    expenses.series[0].graphicalProperties.solidFill = 'C451A0'
ws.add_chart(expenses, 'C35')

# 4. Top incoming launches. I10:J14 are fed by the dynamic helper formulas Z/AA.
entries = BarChart()
entries.type = 'bar'
entries.style = 10
entries.title = 'Top 5 — principais entradas'
entries.x_axis.title = 'Valor (R$)'
entries.height = 8.0
entries.width = 13.0
entries.add_data(Reference(ws, min_col=10, min_row=9, max_row=14), titles_from_data=True)
entries.set_categories(Reference(ws, min_col=9, min_row=10, max_row=14))
entries.legend = None
entries.dataLabels = DataLabelList()
entries.dataLabels.showVal = True
if entries.series:
    entries.series[0].graphicalProperties.solidFill = '35B8C9'
ws.add_chart(entries, 'M35')

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
except AttributeError:
    pass

wb.save(output)
print(output)
print('DASHBOARD_CHARTS', len(ws._charts))
print('UPDATED_FORMULAS', ws['E5'].value, ws['I5'].value, ws['M5'].value)
